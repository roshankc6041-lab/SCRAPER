"""
========================================================
  Kamana Sewa Bank Scraper → Excel → Gmail
  Direct API use garnxa - sabai branches ek paltai!
  
  SETUP:
  pip install requests openpyxl google-auth-oauthlib google-auth-httplib2 google-api-python-client
========================================================
"""

import os, sys, base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

for pkg in ["requests", "openpyxl", "google-auth-oauthlib", "google-auth-httplib2", "google-api-python-client"]:
    try:
        __import__(pkg.replace("-","_").split("-")[0])
    except ImportError:
        import subprocess
        print(f"Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CONFIG = {
    "recipient_email": "roshankc6041@gmail.com",
    "email_subject": "Kamana Sewa Bank - Sabai Branch Data",
    "gmail_credentials_file": "credentials.json",
    "gmail_token_file": "token.json",
}

GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]


def fetch_branches():
    """Direct API bata sabai branches"""
    api_url = "https://backend.kamanasewabank.com/api/v1/branch/map"
    print(f"\n🌐 API call: {api_url}")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/148.0.0.0 Safari/537.36",
        "Referer": "https://www.kamanasewabank.com/",
        "Accept": "application/json",
        "Apikey": "28Q3RB1VNW05EG7T",
    }
    resp = requests.get(api_url, headers=headers, timeout=30)
    resp.raise_for_status()
    
    # Response text bata JSON parse garnus
    text = resp.text.strip()
    if not text:
        raise Exception("Empty response from API")
    
    import json
    data = json.loads(text)
    
    # "data" key bata branches lhinchhha
    if isinstance(data, dict) and "data" in data:
        branches = data["data"]
    elif isinstance(data, list):
        branches = data
    else:
        branches = [data]
    
    print(f"   ✅ {len(branches)} branches milyo!")
    return branches


def create_excel(branches, filename):
    print(f"\n📊 Excel banaudai: {filename}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Branch Data"

    header_fill = PatternFill("solid", start_color="003366")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Clean headers — nested objects flatten garnus
    headers = ["SN", "Branch Name", "Address", "Province", "District", "Phone", "Mobile", "Email", "Latitude", "Longitude", "Map URL"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1,1).value = "Kamana Sewa Bikas Bank - Branch Locations"
    ws.cell(1,1).font = Font(bold=True, size=14, name="Arial", color="003366")
    ws.cell(1,1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2,1).value = f"Total Branches: {len(branches)}  |  Source: kamanasewabank.com  |  Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.cell(2,1).font = Font(italic=True, size=9, name="Arial", color="666666")
    ws.cell(2,1).alignment = Alignment(horizontal="center")
    ws.append([])

    # Headers row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[4].height = 22

    # Data rows
    for i, b in enumerate(branches):
        ws_row = i + 5
        province = b.get("province", {})
        province_name = province.get("title", "") if isinstance(province, dict) else str(province)
        district = b.get("district", {})
        district_name = district.get("dist_title", "") if isinstance(district, dict) else str(district)

        row = [
            i + 1,
            b.get("branch_title", ""),
            b.get("address", ""),
            province_name,
            district_name,
            b.get("phone", ""),
            b.get("mobile", ""),
            b.get("email", ""),
            b.get("lat", ""),
            b.get("lng", ""),
            b.get("map_url", ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(ws_row, col, val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ws_row % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EEF4FB")
        ws.row_dimensions[ws_row].height = 18

    # Column widths
    widths = [5, 28, 35, 20, 18, 20, 15, 32, 12, 12, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filename)
    print(f"   ✅ Excel saved: {filename}")
    return filename


def get_gmail_service():
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(CONFIG["gmail_token_file"]):
        creds = Credentials.from_authorized_user_file(CONFIG["gmail_token_file"], GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CONFIG["gmail_credentials_file"]):
                print(f"\n❌ '{CONFIG['gmail_credentials_file']}' file xaina!")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(CONFIG["gmail_credentials_file"], GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(CONFIG["gmail_token_file"], "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def send_email(service, recipient, subject, excel_file, branch_count):
    print(f"\n📧 Mail pathaudai: {recipient}")
    body = f"""Namaskar,

Kamana Sewa Bank ko sabai {branch_count} branches ko data extract garera Excel maa tayar gariyeko cha.

Excel maa yo information xa:
- Branch Name
- Address
- Province & District
- Phone & Mobile
- Email
- GPS Coordinates (Lat/Lng)
- Google Map URL

Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')}

Excel file attach gariyeko cha.

Dhanyabad!
"""
    msg = MIMEMultipart()
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(excel_file, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(excel_file)}")
        msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    print(f"   ✅ Mail pathiyeko cha: {recipient}")


def main():
    print("=" * 55)
    print("  Kamana Sewa Bank Scraper → Excel → Gmail")
    print("=" * 55)

    recipient = CONFIG["recipient_email"]

    # API bata data
    try:
        branches = fetch_branches()
    except Exception as e:
        print(f"\n❌ API error: {e}")
        sys.exit(1)

    # Excel banaunus
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"kamana_bank_branches_{timestamp}.xlsx"
    create_excel(branches, excel_file)

    # Gmail
    try:
        service = get_gmail_service()
        send_email(service, recipient, CONFIG["email_subject"], excel_file, len(branches))
    except Exception as e:
        print(f"\n❌ Gmail error: {e}")
        print("   Excel file chai bani sakyeko cha, manually pathaunus.")

    print(f"\n{'='*55}")
    print(f"  ✅ Kaam siddhiyo!")
    print(f"  📁 Excel: {excel_file}  ({len(branches)} branches)")
    print(f"  📧 Mail: {recipient}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
